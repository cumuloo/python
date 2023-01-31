import openpyxl
import json
from TikTokApi import TikTokApi

with TikTokApi() as api:

    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = 'No'
    ws['B1'] = 'Username'
    ws['C1'] = 'Follower'
    ws['D1'] = 'Following'
    ws['E1'] = 'Signature'

    no = 1
    for video in api.hashtag(name='bbmnaik').videos():
        video = video.info_full()
        #print(video)

        nickname = (video['itemInfo']['itemStruct']['author']['nickname'])
        signature = (video['itemInfo']['itemStruct']['author']['signature'])
        follower = (video['itemInfo']['itemStruct']['authorStats']['followerCount'])
        following = (video['itemInfo']['itemStruct']['authorStats']['followingCount'])

        x = no + 1
        numCell = str(x)
        ws['A' + numCell] = no
        ws['B' + numCell] = nickname
        ws['C' + numCell] = follower
        ws['D' + numCell] = following
        ws['E' + numCell] = signature

        print(str(no)+ ". " + nickname + ", " + signature + ", follower " + str(follower) + ", following " + str(following))
        no = no + 1

    wb.save('contoh.xlsx')